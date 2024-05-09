import openpyxl.workbook
import streamlit as st
import pandas as pd
import io
import convertePlanilha
import openpyxl
from datetime import datetime

def convert_df(df):
    # Create a temporary workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    for col_index, col in enumerate(df.columns):
        ws.cell(row=1, column=col_index+1).value = col

    # Write data rows
    for index, row in df.iterrows():
        for col_index, value in enumerate(row):
            ws.cell(row=index+2, column=col_index+1).value = value

    # Create a byte stream of the XLSX file in memory
    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_buffer.seek(0)  # Rewind the buffer to the beginning

    return xlsx_buffer.getvalue()

with st.container():

    upload_file = st.file_uploader("Escolha o arquivo", help="xls, xlsx")

    if upload_file is not None:

        df = pd.read_excel(upload_file)
        df_clean = convertePlanilha.clean_data(df.copy())
        st.write(df_clean)
        xlsx_data = convert_df(df_clean)
        
        st.download_button(
            label="Downlaod da planilha tratada em formato xlsx",
            data=xlsx_data,
            file_name=f"{datetime.now()}-{upload_file.name}",
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )