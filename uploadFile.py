import pandas as pd
import openpyxl
import io

def convert_df(df):
    # Create a temporary workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    for col_index, col in enumerate(df.columns):
        ws.cell(row=1, column=col_index+1).value = col

    # Write data rows
    for index, row in df.iterrows():
        for col_index, value in enumerate(row):
            ws.cell(row=index+2, column=col_index+1).value = value

    # Create a byte stream of the XLSX file in memory
    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_buffer.seek(0)  # Rewind the buffer to the beginning

    return xlsx_buffer.getvalue()

def convert_df_to_csv(df:pd.DataFrame):
    return df.to_csv(index=False).encode("utf-8")